"""
Generates a downloadable Excel file replicating the unit mix section (E1:N28)
of the analysis template.
"""

from __future__ import annotations

from io import BytesIO

import openpyxl
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side, numbers
)
from openpyxl.utils import get_column_letter

from app.analyzer import UnitMixSummary, UnitTypeRow


# ── palette ──────────────────────────────────────────────────────────────────
DARK_BLUE  = "1F2D4E"
MID_BLUE   = "2E4272"
LIGHT_BLUE = "D6E4F7"
ACCENT     = "4472C4"
TOTAL_GREY = "F2F2F2"
WHITE      = "FFFFFF"

THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _header_font(bold=True, color=WHITE, size=10):
    return Font(name="Calibri", bold=bold, color=color, size=size)

def _body_font(bold=False, color="000000", size=10):
    return Font(name="Calibri", bold=bold, color=color, size=size)

def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _right() -> Alignment:
    return Alignment(horizontal="right", vertical="center")


CURRENCY_FMT = '"$"#,##0.00'
PCT_FMT      = "0.00%"
NUMBER_FMT   = "#,##0"


def export_unit_mix(summary: UnitMixSummary) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Unit Mix"

    # ── column widths ────────────────────────────────────────────────────────
    col_widths = {
        "A": 16,   # Unit Type (SF)
        "B": 11,   # # of Units
        "C": 14,   # # of Occupied
        "D": 13,   # # of Vacant
        "E": 8,    # %
        "F": 14,   # GPR
        "G": 14,   # Loss to Lease
        "H": 14,   # In Place Rent
        "I": 13,   # Vacancy
        "J": 18,   # Net Rental Income
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 36

    # ── row 1: title ─────────────────────────────────────────────────────────
    ws.merge_cells("A1:J1")
    title_cell = ws["A1"]
    title_cell.value = "Unit Mix"
    if summary.property_name:
        title_cell.value = f"Unit Mix — {summary.property_name}"
    if summary.as_of:
        title_cell.value += f"  (As Of {summary.as_of})"
    title_cell.font = Font(name="Calibri", bold=True, size=13, color=WHITE)
    title_cell.fill = _fill(DARK_BLUE)
    title_cell.alignment = _center()

    # ── row 2: column headers ─────────────────────────────────────────────────
    headers = [
        "Unit Type\n(SF)",
        "# of\nUnits",
        "# of\nOccupied\nUnits",
        "# of\nVacant\nUnits",
        "%",
        "GPR",
        "Loss to\nLease",
        "In Place\nRent",
        "Vacancy",
        "Net Rental\nIncome",
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.font = _header_font()
        cell.fill = _fill(MID_BLUE)
        cell.alignment = _center()
        cell.border = BORDER

    # ── data rows ─────────────────────────────────────────────────────────────
    for i, row in enumerate(summary.rows):
        r = i + 3
        ws.row_dimensions[r].height = 18
        fill = _fill(LIGHT_BLUE) if i % 2 == 0 else _fill(WHITE)

        values = [
            (row.sqft,            NUMBER_FMT,   _center()),
            (row.num_units,       NUMBER_FMT,   _center()),
            (row.num_occupied,    NUMBER_FMT,   _center()),
            (row.num_vacant,      NUMBER_FMT,   _center()),
            (row.pct,             PCT_FMT,      _center()),
            (row.gpr,             CURRENCY_FMT, _right()),
            (row.loss_to_lease,   CURRENCY_FMT, _right()),
            (row.in_place_rent,   CURRENCY_FMT, _right()),
            (row.vacancy,         CURRENCY_FMT, _right()),
            (row.net_rental_income, CURRENCY_FMT, _right()),
        ]
        for col, (val, fmt, align) in enumerate(values, start=1):
            cell = ws.cell(row=r, column=col, value=round(val, 4))
            cell.number_format = fmt
            cell.alignment = align
            cell.font = _body_font()
            cell.fill = fill
            cell.border = BORDER

    # ── spacer row ────────────────────────────────────────────────────────────
    spacer_row = len(summary.rows) + 3
    ws.row_dimensions[spacer_row].height = 6

    # ── totals row ────────────────────────────────────────────────────────────
    totals_row = spacer_row + 1
    ws.row_dimensions[totals_row].height = 18

    totals_label = ws.cell(row=totals_row, column=1, value="Totals")
    totals_label.font = _header_font(color="000000")
    totals_label.fill = _fill(TOTAL_GREY)
    totals_label.alignment = _center()
    totals_label.border = BORDER

    totals_data = [
        (summary.total_units,    NUMBER_FMT,   2),
        (summary.total_occupied, NUMBER_FMT,   3),
        (summary.total_vacant,   NUMBER_FMT,   4),
    ]
    for val, fmt, col in totals_data:
        cell = ws.cell(row=totals_row, column=col, value=val)
        cell.number_format = fmt
        cell.font = _header_font(color="000000")
        cell.fill = _fill(TOTAL_GREY)
        cell.alignment = _center()
        cell.border = BORDER

    # blank cols 5-10 in totals
    for col in range(5, 11):
        cell = ws.cell(row=totals_row, column=col)
        cell.fill = _fill(TOTAL_GREY)
        cell.border = BORDER

    # ── monthly / annual summary rows ────────────────────────────────────────
    def _summary_row(r: int, label: str, gpr, loss, in_place, vacancy, net):
        ws.row_dimensions[r].height = 18
        # cols 1-4 empty but styled
        for col in range(1, 5):
            cell = ws.cell(row=r, column=col)
            cell.fill = _fill(DARK_BLUE)
            cell.border = BORDER

        ws.cell(row=r, column=5, value=label).font = _header_font()
        ws["E" + str(r)].fill = _fill(DARK_BLUE)
        ws["E" + str(r)].alignment = _center()
        ws["E" + str(r)].border = BORDER

        for col, val in enumerate([gpr, loss, in_place, vacancy, net], start=6):
            cell = ws.cell(row=r, column=col, value=round(val, 2))
            cell.number_format = CURRENCY_FMT
            cell.font = _header_font()
            cell.fill = _fill(DARK_BLUE)
            cell.alignment = _right()
            cell.border = BORDER

    monthly_row = totals_row + 1
    annual_row  = totals_row + 2

    _summary_row(
        monthly_row, "Monthly",
        summary.monthly_gpr, summary.monthly_loss,
        summary.monthly_in_place, summary.monthly_vacancy, summary.monthly_net,
    )
    _summary_row(
        annual_row, "Annually",
        summary.annual_gpr, summary.annual_loss,
        summary.annual_in_place, summary.annual_vacancy, summary.annual_net,
    )

    # ── freeze panes ──────────────────────────────────────────────────────────
    ws.freeze_panes = "A3"

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
